import openpyxl, datetime, pyperclip
from openpyxl import Workbook, load_workbook
from datetime import timedelta, date

filepath = 'files/'
filename = 'Timesheet.xlsx'
wb = openpyxl.load_workbook(filepath + filename)
ws = wb.active
ws.protection.sheet = False

def number_to_string_month(string):
    m = {
        '01': 'Jan',
        '02': 'Feb',
        '03': 'Mar',
        '04': 'Apr',
        '05': 'May',
        '06': 'Jun',
        '07': 'Jul',
        '08': 'Aug',
        '09': 'Sep',
        '10': 'Oct',
        '11': 'Nov',
        '12': 'Dec'
        }
    try:
        out = m[string]
        return out
    except:
        raise ValueError('Not a month')

def SetWorkWeek(LastEndingDate):
    date2 = LastEndingDate.split('-')
    intdate2 = [int(item) for item in date2]
    date3 = date(intdate2[0],intdate2[1],intdate2[2])
    endDate = date3 + timedelta(days=1)
    # datetime = str(endDate) + ' 00:00:00'
    datetime = str(endDate)
    ws['C30'].value = datetime
    return date3


def ClearOverTime(OTamt):
    i = 60
    numberOT = int(OTamt) + 59
    columns = ['A','B','C','E','H','J']
    while i < numberOT:
        for k in columns:
            cell = k + str(i)
            ws[cell].value = None
        i += 1

def SetOverTime(instance):
    columns = ['A', 'B', 'C', 'E', 'H', 'J']
    row = 60 + instance
    while True:
        date = input('Enter OT Date: ')
        if date == '':
            break
        dateCell = columns[0] + str(row)
        ws[dateCell].value = date
        print(ws[dateCell].value)

        startTime = input('Enter Start Time 24hr clock: ')
        if startTime == '':
            break
        startTimeCell = columns[1] + str(row)
        ws[startTimeCell].value = startTime
        print(ws[startTimeCell].value)

        endTime = input('Enter End Time 24hr clock: ')
        if endTime == '':
            break
        endTimeCell = columns[2] + str(row)
        ws[endTimeCell].value = endTime
        print(ws[endTimeCell].value)

        SR = input('Enter SR#: ')
        if SR == '':
            break
        typeCell = columns[3] + str(row)
        ws[typeCell].value = 'OT (+)'
        srCell = columns[4] + str(row)
        ws[srCell].value = SR
        print(ws[srCell].value)

        Description = input('Enter Description: ')
        if Description == '':
            break
        descrCell = columns[5] + str(row)
        ws[descrCell].value = Description
        print(ws[descrCell].value)

        finished = input('Done? yes or no: ')
        if finished == 'yes' or 'no':
            break

def Save_File_Date(StartingPeriod):
    FileDate = StartingPeriod + timedelta(days=14)
    FileDate2 = str(FileDate).split('-')
    filename = '.OTINF Kenneth Situ ' + number_to_string_month(str(FileDate2[1])) + ' ' + str(FileDate2[2]) + ', ' + str(FileDate2[0]) + '.xlsx'
    wb.save(filepath + filename)
    return filename

if __name__ == '__main__':
    LastPayPeriod = input("Enter last Pay Period End Date (yyyy-mm-dd): ")
    StartingPeriod = SetWorkWeek(LastPayPeriod)
    # ClearOverTime(15)
    i = 0
    while True:
        checkOT = input('Input OT? yes or no: ')
        if checkOT == 'no':
            break
        else:
            SetOverTime(i)
            i += 1

    SavedFile = Save_File_Date(StartingPeriod)
    pyperclip.copy(SavedFile)
    print(SavedFile)
